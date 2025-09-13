import os
import subprocess
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

BASE_DIR = os.getcwd()

folders = [
    "1. Player Details",
    "2. Player List",
    "3. Search Player",
    "4. Top Players",
    "5. Top Team"
]

wb = Workbook()
ws = wb.active

ws.merge_cells("A1:J1")

cell = ws["A1"]
cell.value = "TEST RESULTS"
cell.alignment = Alignment(horizontal="center", vertical="center")
cell.font = Font(bold=True, size=20)

ws.append(["ITEM", "ID", "RESULT", "STATUS"])
for cell in ws[2]:
    cell.font = Font(bold=True)

if not os.path.exists('res'):
    os.mkdir('res')

wb.save("res/test_results.xlsx")

for folder in folders:
    path = os.path.join(BASE_DIR, folder)
    if not os.path.exists(path):
        print(f"[WARNING] Folder not found: {path}")
        continue

    for file in sorted(os.listdir(path)):
        if file.endswith(".py"):
            script_path = os.path.join(path, file)
            print(f"\n Executing: {script_path}")
            subprocess.run(["python", script_path])

wb = load_workbook("res/test_results.xlsx")
ws = wb.active

for col in ws.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    ws.column_dimensions[col_letter].width = max_length + 2

wb.save("res/test_results.xlsx")
